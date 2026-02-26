"""
Xlsx Skill - Microsoft Excel Spreadsheet Operations
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class XlsxSkill:
    name = "xlsx"
    description = "Create and edit Excel spreadsheets"
    
    def create_xlsx(self, filename, sheet_name="Sheet"):
        """Create a new Excel workbook"""
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(sheet_name)
        wb.save(filename)
        return {"result": "created", "file": filename}
    
    def add_sheet(self, filename, sheet_name):
        """Add a new sheet"""
        wb = load_workbook(filename)
        wb.create_sheet(sheet_name)
        wb.save(filename)
        return {"result": "added", "sheet": sheet_name}
    
    def add_data(self, filename, sheet_name, data):
        """Add data to sheet"""
        wb = load_workbook(filename)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, cell_data in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_data)
        
        wb.save(filename)
        return {"result": "added", "rows": len(data)}
    
    def add_formula(self, filename, sheet_name, cell, formula):
        """Add formula to cell"""
        wb = load_workbook(filename)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws[cell] = formula
        wb.save(filename)
        return {"result": "added", "cell": cell, "formula": formula}
    
    def add_style(self, filename, sheet_name, cell_range, bold=True, fill_color="FFFF00"):
        """Style cells"""
        wb = load_workbook(filename)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        font = Font(bold=bold)
        
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = fill
                cell.font = font
        
        wb.save(filename)
        return {"result": "styled", "range": cell_range}

skill = XlsxSkill()
